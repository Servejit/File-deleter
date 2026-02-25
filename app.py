import streamlit as st
import os
import zipfile
from io import BytesIO
from openpyxl import load_workbook

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="Excel Auto Processor",
    layout="centered"
)

st.title("📊 Excel Auto Processor System")


# =====================================================
# FILE MAP
# =====================================================

file_map = {
    "abc.xlsx": "ABC.xlsx",
    "xyz.xlsx": "XYZ.xlsx",
    "pqr.xlsx": "PQR.xlsx"
}


# =====================================================
# MOVE RULES
# =====================================================

moves = [
    ("qYY", "r"),
    ("qY", "qYY"),
    ("q", "qY"),
    ("pYY", "q"),
    ("pY", "pYY"),
    ("Y", "pY")
]


# =====================================================
# STEP 1 DELETE FILES
# =====================================================

st.header("Step 1: Delete Old Files")

if st.button("🗑 Delete Old Files"):

    deleted = False

    for file in list(file_map.keys()) + list(file_map.values()):

        path = os.path.join(os.getcwd(), file)

        if os.path.exists(path):

            os.remove(path)

            st.success(f"Deleted: {file}")

            deleted = True

    if not deleted:

        st.info("No old files found")


# =====================================================
# STEP 2 UPLOAD FILES (FIXED)
# =====================================================

st.header("Step 2: Upload New Files")

uploaded_files = st.file_uploader(
    "Upload abc.xlsx, xyz.xlsx, pqr.xlsx",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded_files:

    for uploaded_file in uploaded_files:

        try:

            save_path = os.path.join(os.getcwd(), uploaded_file.name)

            with open(save_path, "wb") as f:

                f.write(uploaded_file.getvalue())

            st.success(f"✅ Uploaded: {uploaded_file.name}")

        except Exception as e:

            st.error(f"❌ Upload failed: {uploaded_file.name}")
            st.error(e)


# =====================================================
# STEP 3 PROCESS FILES
# =====================================================

st.header("Step 3: Process and Download")

if st.button("⚙ Process Files and Prepare Download"):

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w") as zip_file:

        for input_file, output_file in file_map.items():

            input_path = os.path.join(os.getcwd(), input_file)

            if not os.path.exists(input_path):

                st.error(f"❌ Missing file: {input_file}")

                continue

            wb = load_workbook(input_path)

            data_cache = {}

            # READ DATA

            for src, dst in moves:

                if src in wb.sheetnames:

                    sheet = wb[src]

                    data_cache[src] = [

                        [sheet.cell(row=r, column=c).value for c in range(1, 16)]

                        for r in range(2, 52)

                    ]

            # WRITE DATA

            for src, dst in moves:

                if dst in wb.sheetnames and src in data_cache:

                    sheet = wb[dst]

                    for r_idx, row in enumerate(data_cache[src], start=2):

                        for c_idx, val in enumerate(row, start=1):

                            sheet.cell(row=r_idx, column=c_idx).value = val


            # SAVE TO MEMORY

            file_buffer = BytesIO()

            wb.save(file_buffer)


            zip_file.writestr(

                output_file,

                file_buffer.getvalue()

            )


            st.success(f"✅ Processed: {output_file}")


    zip_buffer.seek(0)


    st.download_button(

        label="⬇ Download ALL Files",

        data=zip_buffer,

        file_name="Processed_Excel_Files.zip",

        mime="application/zip"

    )


# =====================================================
# REFRESH BUTTON
# =====================================================

st.header("Step 4: Refresh")

if st.button("🔄 Refresh App"):

    st.rerun()
